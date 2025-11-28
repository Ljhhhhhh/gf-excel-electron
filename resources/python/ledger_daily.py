#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Daily ledger updater for multiple sheets:
- "融资及还款明细" sheet (docs/1融资及还款明细说明.txt)
- "资产明细" sheet (docs/2资产明细说明.txt)

Rules:
- preserve every existing cell (values, formulas, styles);
- append new rows based on target date;
- apply specific formatting and transformations per sheet.
"""

from __future__ import annotations

import argparse
import datetime as dt
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.styles import PatternFill, Font, Alignment, alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.datetime import from_excel

# === 常量 ===

TEMPLATE_ROW_INDEX = 10
SHEET_FINANCING_REPAYMENT = "融资及还款明细"
SHEET_ASSET_DETAIL = "资产明细"
SHEET_ZHONGDENG = "中登登记表"
SHEET_CUSTOMER = "客户表"
SHEET_INTEREST = "利息缴纳"
CUSTOMER_SOURCE_SHEET = "sheet1"

# 目标表列
COL_A = column_index_from_string("A")
COL_B = column_index_from_string("B")
COL_C = column_index_from_string("C")
COL_D = column_index_from_string("D")
COL_E = column_index_from_string("E")
COL_F = column_index_from_string("F")
COL_G = column_index_from_string("G")
COL_H = column_index_from_string("H")
COL_I = column_index_from_string("I")
COL_J = column_index_from_string("J")
COL_K = column_index_from_string("K")
COL_L = column_index_from_string("L")
COL_M = column_index_from_string("M")
COL_N = column_index_from_string("N")
COL_O = column_index_from_string("O")
COL_P = column_index_from_string("P")
COL_Q = column_index_from_string("Q")
COL_R = column_index_from_string("R")
COL_S = column_index_from_string("S")
COL_T = column_index_from_string("T")
COL_U = column_index_from_string("U")
COL_V = column_index_from_string("V")
COL_W = column_index_from_string("W")
COL_X = column_index_from_string("X")
COL_Y = column_index_from_string("Y")
COL_Z = column_index_from_string("Z")
COL_AA = column_index_from_string("AA")
COL_AB = column_index_from_string("AB")
COL_AC = column_index_from_string("AC")
COL_AD = column_index_from_string("AD")
COL_AE = column_index_from_string("AE")
COL_AF = column_index_from_string("AF")
COL_AG = column_index_from_string("AG")
COL_AH = column_index_from_string("AH")
COL_AI = column_index_from_string("AI")
COL_AJ = column_index_from_string("AJ")
COL_AK = column_index_from_string("AK")
COL_AO = column_index_from_string("AO")
COL_AP = column_index_from_string("AP")
COL_AQ = column_index_from_string("AQ")
COL_AR = column_index_from_string("AR")

# 放款明细列
LOAN_COL_B = column_index_from_string("B")
LOAN_COL_AE = column_index_from_string("AE")
LOAN_COL_AG = column_index_from_string("AG")
LOAN_COL_K = column_index_from_string("K")
LOAN_COL_C = column_index_from_string("C")
LOAN_COL_G = column_index_from_string("G")
LOAN_COL_AZ = column_index_from_string("AZ")
LOAN_COL_J = column_index_from_string("J")
LOAN_COL_AA = column_index_from_string("AA")
LOAN_COL_L = column_index_from_string("L")
LOAN_COL_AK = column_index_from_string("AK")
LOAN_COL_N = column_index_from_string("N")
LOAN_COL_M = column_index_from_string("M")
LOAN_COL_Y = column_index_from_string("Y")
LOAN_COL_AW = column_index_from_string("AW")
LOAN_COL_P = column_index_from_string("P")
LOAN_COL_BC = column_index_from_string("BC")
LOAN_COL_BF = column_index_from_string("BF")
LOAN_COL_Q = column_index_from_string("Q")
LOAN_COL_T = column_index_from_string("T")
LOAN_COL_U = column_index_from_string("U")

# 还款明细列（保理/再保理）
REPAY_COL_B = column_index_from_string("B")
REPAY_COL_C = column_index_from_string("C")
REPAY_COL_F = column_index_from_string("F")
REPAY_COL_G = column_index_from_string("G")
REPAY_COL_H = column_index_from_string("H")
REPAY_COL_J = column_index_from_string("J")
REPAY_COL_M = column_index_from_string("M")
REPAY_COL_O = column_index_from_string("O")
REPAY_COL_X = column_index_from_string("X")
REPAY_COL_Y = column_index_from_string("Y")
REPAY_COL_AB = column_index_from_string("AB")
REPAY_COL_AC = column_index_from_string("AC")
REPAY_COL_AD = column_index_from_string("AD")
REPAY_COL_AE = column_index_from_string("AE")
REPAY_COL_AG = column_index_from_string("AG")
REPAY_COL_AH = column_index_from_string("AH")

# 放款明细列（资产明细用）- 补充未定义的列
LOAN_COL_D = column_index_from_string("D")
LOAN_COL_E = column_index_from_string("E")
LOAN_COL_F = column_index_from_string("F")
LOAN_COL_S = column_index_from_string("S")
LOAN_COL_AC = column_index_from_string("AC")
LOAN_COL_AF = column_index_from_string("AF")
LOAN_COL_AH = column_index_from_string("AH")
LOAN_COL_AI = column_index_from_string("AI")
LOAN_COL_AJ = column_index_from_string("AJ")
LOAN_COL_AL = column_index_from_string("AL")
LOAN_COL_AM = column_index_from_string("AM")
LOAN_COL_AN = column_index_from_string("AN")
LOAN_COL_AO = column_index_from_string("AO")
LOAN_COL_AQ = column_index_from_string("AQ")
LOAN_COL_AR = column_index_from_string("AR")

# 涓櫥鐧昏琛ㄧ澶达細婧愭枃浠朵腑鐨勫垪鎸囧畾
ZD_COL_C = column_index_from_string("C")
ZD_COL_D = column_index_from_string("D")
ZD_COL_E = column_index_from_string("E")
ZD_COL_F = column_index_from_string("F")
ZD_COL_G = column_index_from_string("G")
ZD_COL_H = column_index_from_string("H")
ZD_COL_I = column_index_from_string("I")
ZD_COL_J = column_index_from_string("J")
ZD_COL_K = column_index_from_string("K")
ZD_COL_L = column_index_from_string("L")
ZD_COL_M = column_index_from_string("M")
ZD_COL_N = column_index_from_string("N")
ZD_COL_O = column_index_from_string("O")
ZD_COL_P = column_index_from_string("P")
ZD_COL_R = column_index_from_string("R")
ZD_COL_S = column_index_from_string("S")
ZD_COL_T = column_index_from_string("T")
ZD_COL_U = column_index_from_string("U")
ZD_COL_W = column_index_from_string("W")
ZD_COL_X = column_index_from_string("X")
ZD_COL_Y = column_index_from_string("Y")

# 下载的《客户表》列索引（基于需求文档）
CUSTOMER_SRC_COL_NAME = column_index_from_string("A")
CUSTOMER_SRC_COL_CODE = column_index_from_string("B")
CUSTOMER_SRC_COL_INDUSTRY = column_index_from_string("C")
CUSTOMER_SRC_COL_ECONOMIC = column_index_from_string("D")
CUSTOMER_SRC_COL_SCALE = column_index_from_string("E")
CUSTOMER_SRC_COL_REGISTER_ADDR = column_index_from_string("F")
CUSTOMER_SRC_COL_BUSINESS_ADDR = column_index_from_string("G")
CUSTOMER_SRC_COL_ROLE = column_index_from_string("H")
CUSTOMER_SRC_COL_LEGAL_REP = column_index_from_string("I")
CUSTOMER_SRC_COL_LEGAL_ID = column_index_from_string("J")
CUSTOMER_SRC_COL_REGION = column_index_from_string("L")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Append ledger financing & repayment rows.")
    parser.add_argument("--ledger", required=True, help="现有台账文件路径")
    parser.add_argument("--loan", required=True, help="放款明细路径")
    parser.add_argument("--factoring-repay", required=True, help="保理融资还款明细路径")
    parser.add_argument("--refactoring-repay", required=True, help="再保理融资还款明细路径")
    parser.add_argument("--zhongdeng", required=True, help="中登登记表路径")
    parser.add_argument("--customer", required=True, help="客户表路径（下载版）")
    parser.add_argument("--date", required=True, help="目标日期，格式 YYYYMMDD")
    parser.add_argument("--output", required=True, help="输出文件路径")
    return parser.parse_args()


def parse_input_date(date_str: str) -> dt.date:
    try:
        return dt.datetime.strptime(date_str, "%Y%m%d").date()
    except ValueError as exc:  # pragma: no cover - 防御性校验
        raise SystemExit(f"无效日期格式: {date_str}，需为 YYYYMMDD") from exc


def normalize_excel_date(value) -> Optional[dt.date]:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date()
        except Exception:
            return None
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
            try:
                return dt.datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None


def normalize_string(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (int, float)):
        return str(value).strip()
    return str(value).strip()


def find_sheet_by_name(wb, sheet_name: str) -> "Worksheet":
    """ 根据名称查找工作表 """
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    raise SystemExit(f"未找到目标工作表：{sheet_name}")


def find_last_data_row(ws) -> int:
    # 使用 A 列（序号公式）向上查找
    for row_idx in range(ws.max_row, 0, -1):
        if ws.cell(row=row_idx, column=COL_A).value not in (None, ""):
            return row_idx
    return 0


def get_last_existing_date(ws) -> Optional[dt.date]:
    last_row = find_last_data_row(ws)
    if last_row == 0:
        return None
    last_w = normalize_excel_date(ws.cell(row=last_row, column=COL_W).value)
    last_ae = normalize_excel_date(ws.cell(row=last_row, column=COL_AE).value)
    return last_w or last_ae


def cache_template_row(ws, row_index: int) -> Dict[int, Dict[str, object]]:
    cached: Dict[int, Dict[str, object]] = {}
    for cell in ws[row_index]:
        translator = None
        if cell.data_type == "f" and isinstance(cell.value, str):
            origin = f"{get_column_letter(cell.column)}{row_index}"
            translator = Translator(cell.value, origin=origin)
        cached[cell.column] = {
            "value": cell.value,
            "data_type": cell.data_type,
            # 样式对象可共享，避免为每个单元格复制带来的样式爆炸与慢速保存
            "style": cell._style,
            "translator": translator,
        }
    return cached


def apply_template_row(ws, template_cache, target_row: int, template_row_index: int, template_height: Optional[float]):
    for col_idx, meta in template_cache.items():
        target_cell = ws.cell(row=target_row, column=col_idx)
        tpl_value = meta["value"]
        tpl_type = meta["data_type"]
        translator: Optional[Translator] = meta.get("translator")  # type: ignore
        if meta.get("style"):
            target_cell._style = meta["style"]

        if tpl_type == "f" and translator:
            target = f"{get_column_letter(col_idx)}{target_row}"
            # openpyxl 3.1 的接口是 translate_formula，而旧版本是 translate
            if hasattr(translator, "translate_formula"):
                target_cell.value = translator.translate_formula(target)
            else:  # pragma: no cover - 兼容旧接口
                target_cell.value = translator.translate(target)
        else:
            target_cell.value = tpl_value

    if template_height:
        ws.row_dimensions[target_row].height = template_height


def apply_b_column_conditional_format(ws, start_row: int, end_row: int):
    """
    对 B 列指定范围应用条件格式：
    - 当公式计算结果为文本时，应用黄底红字加粗样式
    - 当公式计算结果为数字时，不应用样式（使用模板默认样式）
    """
    if start_row > end_row:
        return

    # 定义文本时的样式
    text_fill = PatternFill(start_color='FFFFE699', end_color='FFFFE699', fill_type='solid')
    text_font = Font(color='FFC00000', size=11, bold=True)

    # 条件格式规则：当 B 列值为文本时应用样式
    # 使用 ISTEXT 函数判断，$B 表示列固定，行号相对
    formula_rule = FormulaRule(
        formula=[f'ISTEXT($B{start_row})'],
        fill=text_fill,
        font=text_font
    )

    # 应用到 B 列指定范围
    range_string = f'B{start_row}:B{end_row}'
    ws.conditional_formatting.add(range_string, formula_rule)


def collect_loan_rows(path: Path, target_date: dt.date) -> List[Sequence]:
    wb = load_workbook(path, read_only=True, data_only=False)
    ws = wb.active
    matched: List[Sequence] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if normalize_excel_date(row[LOAN_COL_P - 1]) == target_date:
            matched.append(row)
    wb.close()
    return matched


def collect_repay_rows(path: Path, target_date: dt.date, fee_type: str = "本金") -> List[Sequence]:
    wb = load_workbook(path, read_only=True, data_only=False)
    ws = wb.active
    matched: List[Sequence] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if normalize_excel_date(row[REPAY_COL_AE - 1]) != target_date:
            continue
        fee_value = row[REPAY_COL_AB - 1]
        if (fee_value or "").strip() != fee_type:
            continue
        matched.append(row)
    wb.close()
    # 按交易银行流水号（AH）升序
    return sorted(matched, key=lambda r: (r[REPAY_COL_AH - 1] is None, str(r[REPAY_COL_AH - 1])))


def collect_zhongdeng_rows(path: Path, finance_codes: set[str]) -> List[Sequence]:
    if not finance_codes:
        return []

    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        ws = wb[SHEET_ZHONGDENG] if SHEET_ZHONGDENG in wb.sheetnames else wb.active
        dedup: Dict[str, Sequence] = {}
        fallback_index = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            finance_code = normalize_string(row[ZD_COL_C - 1])
            if not finance_code or finance_code not in finance_codes:
                continue

            reg_number = normalize_string(row[ZD_COL_I - 1])
            reg_type = normalize_string(row[ZD_COL_F - 1])
            key = reg_number or f"__row_{fallback_index}"
            if not reg_number:
                fallback_index += 1

            existing = dedup.get(key)
            if existing:
                existing_type = normalize_string(existing[ZD_COL_F - 1])
                if existing_type != "初始登记" and reg_type == "初始登记":
                    dedup[key] = row
                continue

            dedup[key] = row

        return [
            row
            for row in dedup.values()
            if normalize_string(row[ZD_COL_F - 1]) == "初始登记"
        ]
    finally:
        wb.close()


def set_cell(ws, row_idx: int, col_idx: int, value):
    cell = ws.cell(row=row_idx, column=col_idx)
    if value is None:
        cell.value = None
    else:
        cell.value = value


def append_loan_block(ws, template_cache, template_height, template_row_index, rows: Iterable[Sequence]) -> int:
    added = 0
    start_row = ws.max_row
    for record in rows:
        target_row = start_row + added + 1
        apply_template_row(ws, template_cache, target_row, template_row_index, template_height)
        # B 列样式通过条件格式统一处理，不再逐行应用

        set_cell(ws, target_row, COL_C, None)
        set_cell(ws, target_row, COL_D, record[LOAN_COL_B - 1])
        loan_e = record[LOAN_COL_AE - 1]
        if isinstance(loan_e, str):
            loan_e = loan_e.replace('业务', '')

        set_cell(ws, target_row, COL_E, loan_e)
        # F列：明保->明保理，暗保->暗保理
        loan_f = record[LOAN_COL_AG - 1]
        if isinstance(loan_f, str):
            if loan_f.strip() == '明保':
                loan_f = '明保理'
            elif loan_f.strip() == '暗保':
                loan_f = '暗保理'
        set_cell(ws, target_row, COL_F, loan_f)
        set_cell(ws, target_row, COL_G, record[LOAN_COL_K - 1])
        set_cell(ws, target_row, COL_H, record[LOAN_COL_C - 1])
        set_cell(ws, target_row, COL_I, record[LOAN_COL_G - 1])
        loan_j = record[LOAN_COL_AZ - 1]
        loan_k = record[LOAN_COL_J - 1]
        set_cell(ws, target_row, COL_J, loan_j if loan_j not in (None, "") else "/")
        set_cell(ws, target_row, COL_K, loan_k if loan_k not in (None, "") else "/")
        set_cell(ws, target_row, COL_L, record[LOAN_COL_AA - 1])
        set_cell(ws, target_row, COL_N, record[LOAN_COL_AK - 1])
        ws.cell(row=target_row, column=COL_N).alignment = Alignment(horizontal='left')
        # O 列保留模板公式 (=GX&"-"&TX)
        set_cell(ws, target_row, COL_P, record[LOAN_COL_M - 1])
        ws.cell(row=target_row, column=COL_P).alignment = Alignment(horizontal='center')
        set_cell(ws, target_row, COL_Q, record[LOAN_COL_L - 1])
        set_cell(ws, target_row, COL_R, record[LOAN_COL_AK - 1])
        ws.cell(row=target_row, column=COL_R).alignment = Alignment(horizontal='left')
        biz_mode = record[LOAN_COL_Y - 1]
        set_cell(ws, target_row, COL_S, "保理" if isinstance(biz_mode, str) and biz_mode.strip() == "直接投放" else "再保理")
        set_cell(ws, target_row, COL_T, record[LOAN_COL_N - 1])
        set_cell(ws, target_row, COL_U, record[LOAN_COL_AW - 1])
        set_cell(ws, target_row, COL_V, record[LOAN_COL_AW - 1])
        set_cell(ws, target_row, COL_W, record[LOAN_COL_P - 1])
        set_cell(ws, target_row, COL_X, record[LOAN_COL_BC - 1])
        bf_value = record[LOAN_COL_BF - 1]
        if isinstance(bf_value, (int, float)):
            set_cell(ws, target_row, COL_Z, round(float(bf_value), 4))
            ws.cell(row=target_row, column=COL_Z).number_format = "0.00%"
        else:
            set_cell(ws, target_row, COL_Z, bf_value)
        set_cell(ws, target_row, COL_AB, record[LOAN_COL_Q - 1])
        set_cell(ws, target_row, COL_AC, record[LOAN_COL_T - 1])
        set_cell(ws, target_row, COL_AD, record[LOAN_COL_U - 1])
        set_cell(ws, target_row, COL_AO, f"=VLOOKUP(AG{target_row},P:W,8,0)")
        set_cell(ws, target_row, COL_AP, None)
        set_cell(ws, target_row, COL_AQ, f"=XLOOKUP(AG{target_row},P:P,AD:AD)")
        added += 1
    return added


def append_repay_block(ws, template_cache, template_height, template_row_index, rows: Iterable[Sequence], repay_type: str) -> int:
    added = 0
    start_row = ws.max_row
    for record in rows:
        target_row = start_row + added + 1
        apply_template_row(ws, template_cache, target_row, template_row_index, template_height)
        # B 列样式通过条件格式统一处理，不再逐行应用

        set_cell(ws, target_row, COL_C, None)
        set_cell(ws, target_row, COL_D, record[REPAY_COL_G - 1])
        repay_e = record[REPAY_COL_H - 1]
        if isinstance(repay_e, str):
            repay_e = repay_e.replace('业务', '')
        set_cell(ws, target_row, COL_E, repay_e)
        # F列：明保->明保理，暗保->暗保理
        repay_f = record[REPAY_COL_J - 1]
        if isinstance(repay_f, str):
            if repay_f.strip() == '明保':
                repay_f = '明保理'
            elif repay_f.strip() == '暗保':
                repay_f = '暗保理'
        set_cell(ws, target_row, COL_F, repay_f)
        set_cell(ws, target_row, COL_G, record[REPAY_COL_M - 1])
        set_cell(ws, target_row, COL_H, record[REPAY_COL_B - 1])
        set_cell(ws, target_row, COL_I, record[REPAY_COL_C - 1])
        set_cell(ws, target_row, COL_J, "/")
        set_cell(ws, target_row, COL_K, "/")
        set_cell(ws, target_row, COL_L, record[REPAY_COL_F - 1])

        for col_idx in (
            COL_M,
            COL_N,
            COL_O,
            COL_P,
            COL_Q,
            COL_R,
            COL_S,
            COL_T,
            COL_U,
            COL_V,
            COL_W,
            COL_X,
            COL_Y,
            COL_Z,
            COL_AA,
            COL_AB,
            COL_AC,
            COL_AD,
            COL_AP,
        ):
            set_cell(ws, target_row, col_idx, None)

        set_cell(ws, target_row, COL_AE, record[REPAY_COL_AE - 1])
        set_cell(ws, target_row, COL_AG, record[REPAY_COL_O - 1])
        set_cell(ws, target_row, COL_AH, record[REPAY_COL_AG - 1])
        set_cell(ws, target_row, COL_AI, record[REPAY_COL_AG - 1])
        set_cell(ws, target_row, COL_AJ, record[REPAY_COL_AE - 1])
        set_cell(ws, target_row, COL_AK, repay_type)
        set_cell(ws, target_row, COL_AO, f"=VLOOKUP(AG{target_row},P:W,8,0)")
        set_cell(ws, target_row, COL_AQ, f"=XLOOKUP(AG{target_row},P:P,AD:AD)")
        set_cell(ws, target_row, COL_AR, record[REPAY_COL_AH - 1])
        added += 1
    return added


def extract_date_prefix(value):
    text = normalize_string(value)
    if not text:
        return None
    prefix = text.split("~", 1)[0].strip()
    if not prefix:
        return None
    parsed = normalize_excel_date(prefix)
    return parsed or prefix


def append_interest_rows(ws, template_cache, template_height, template_row_index: int, rows: Iterable[Sequence], label: str) -> int:
    added = 0
    start_row = find_last_data_row(ws)

    for record in rows:
        target_row = start_row + added + 1
        apply_template_row(ws, template_cache, target_row, template_row_index, template_height)

        set_cell(ws, target_row, COL_D, record[REPAY_COL_O - 1])
        set_cell(ws, target_row, COL_G, record[REPAY_COL_B - 1])
        set_cell(ws, target_row, COL_J, None)
        set_cell(ws, target_row, COL_K, record[REPAY_COL_AE - 1])
        set_cell(ws, target_row, COL_L, "/")

        prefix_date = extract_date_prefix(record[REPAY_COL_AC - 1])
        set_cell(ws, target_row, COL_M, prefix_date)
        set_cell(ws, target_row, COL_N, prefix_date)
        set_cell(ws, target_row, COL_O, record[REPAY_COL_AD - 1])

        set_cell(ws, target_row, COL_P, None)
        set_cell(ws, target_row, COL_Q, None)
        set_cell(ws, target_row, COL_W, None)

        t_value = record[REPAY_COL_Y - 1]
        if t_value in (None, ""):
            t_value = record[REPAY_COL_X - 1]
        set_cell(ws, target_row, COL_T, t_value)

        amount = record[REPAY_COL_AG - 1]
        set_cell(ws, target_row, COL_U, amount)
        set_cell(ws, target_row, COL_V, amount)

        set_cell(ws, target_row, COL_S, f"=ROUND(U{target_row}*360/T{target_row}/R{target_row},2)")

        added += 1

    if added:
        print(f"[利息缴纳] {label}新增 {added} 行")
    return added


def merge_ai_with_sum(ws, start_row: int, end_row: int, target_date: dt.date):
    if start_row > end_row:
        return

    def finalize(group: List[int], ar_value):
        if not group:
            return
        total = 0.0
        for r in group:
            val = ws.cell(row=r, column=COL_AH).value
            if isinstance(val, (int, float)):
                total += float(val)
            elif val not in (None, ""):
                try:
                    total += float(val)
                except Exception:
                    pass
        top = group[0]
        set_cell(ws, top, COL_AI, total)
        if len(group) > 1:
            ws.merge_cells(start_row=top, end_row=group[-1], start_column=COL_AI, end_column=COL_AI)

    current_group: List[int] = []
    current_ar = None

    for row_idx in range(start_row, end_row + 1):
        ae_date = normalize_excel_date(ws.cell(row=row_idx, column=COL_AE).value)
        ar_value = ws.cell(row=row_idx, column=COL_AR).value
        if ae_date == target_date and ar_value not in (None, ""):
            if current_ar is None:
                current_ar = ar_value
                current_group = [row_idx]
            elif ar_value == current_ar:
                current_group.append(row_idx)
            else:
                finalize(current_group, current_ar)
                current_ar = ar_value
                current_group = [row_idx]
        else:
            finalize(current_group, current_ar)
            current_group = []
            current_ar = None

    finalize(current_group, current_ar)


# =============================================================================
# 资产明细 Sheet 处理函数
# =============================================================================

def append_asset_detail_block(ws, template_cache, template_height, template_row_index: int, rows: Iterable[Sequence]) -> int:
    """
    向【资产明细】sheet 追加放款明细数据
    数据来源：放款明细表，筛选条件为 P 列实际放款日期 = 目标日期
    """
    added = 0
    start_row = ws.max_row
    for record in rows:
        target_row = start_row + added + 1
        apply_template_row(ws, template_cache, target_row, template_row_index, template_height)

        # A列：公式 =ROW()-3
        set_cell(ws, target_row, COL_A, f"=ROW()-3")
        # C列：放款明细 B列 业务来源
        set_cell(ws, target_row, COL_C, record[LOAN_COL_B - 1])
        # D列：固定为"宁波国富商业保理有限公司"
        set_cell(ws, target_row, COL_D, "宁波国富商业保理有限公司")
        # E列：放款明细 AA列 所属行业
        set_cell(ws, target_row, COL_E, record[LOAN_COL_AA - 1])
        # F列：放款明细 E列 经济成分
        set_cell(ws, target_row, COL_F, record[LOAN_COL_E - 1])
        # G列：放款明细 F列 企业规模
        set_cell(ws, target_row, COL_G, record[LOAN_COL_F - 1])
        # H列：放款明细 AI列 产品名称
        set_cell(ws, target_row, COL_H, record[LOAN_COL_AI - 1])
        # I列：放款明细 AC列 是否票据增信
        set_cell(ws, target_row, COL_I, record[LOAN_COL_AC - 1])
        # J列：放款明细 AF列 有/无追索权
        loan_j = record[LOAN_COL_AF - 1]
        if isinstance(loan_j, str):
            if loan_j.strip() == '有追':
                # strip() 方法用于移除字符串开头和结尾的空白字符（包括空格、制表符、换行符等）。
                loan_j = '有追索权'
            elif loan_j.strip() == '无追':
                loan_j = '无追索权'
        set_cell(ws, target_row, COL_J, loan_j)
        # K列：放款明细 AG列 明暗保
        loan_k = record[LOAN_COL_AG - 1]
        if isinstance(loan_k, str):
            if loan_k.strip() == '明保':
                loan_k = '明保理'
            elif loan_k.strip() == '暗保':
                loan_k = '暗保理'
        set_cell(ws, target_row, COL_K, loan_k)
        # M列：放款明细 AJ列 单据类型
        set_cell(ws, target_row, COL_M, record[LOAN_COL_AJ - 1])
        # N列：放款明细 AH列 正反向
        set_cell(ws, target_row, COL_N, record[LOAN_COL_AH - 1])
        # P列：放款明细 C列 保理/再保理申请人名称
        set_cell(ws, target_row, COL_P, record[LOAN_COL_C - 1])
        # Q列：放款明细 D列 统一社会信用代码
        set_cell(ws, target_row, COL_Q, record[LOAN_COL_D - 1])
        # R列：放款明细 G列 基础交易对手方名称（买方/卖方）
        set_cell(ws, target_row, COL_R, record[LOAN_COL_G - 1])
        # S列：放款明细 S列 收票方名称(多个逗号分割)
        set_cell(ws, target_row, COL_S, record[LOAN_COL_S - 1])
        # T列：放款明细 T列 付款方式
        set_cell(ws, target_row, COL_T, record[LOAN_COL_T - 1])
        # U列：放款明细 U列 放款账户开户行
        set_cell(ws, target_row, COL_U, record[LOAN_COL_U - 1])
        # V列：放款明细 K列 资产编号
        set_cell(ws, target_row, COL_V, record[LOAN_COL_K - 1])
        # W列：放款明细 L列 融资申请号
        set_cell(ws, target_row, COL_W, record[LOAN_COL_L - 1])
        # X列：放款明细 L列 融资申请号（同W列）
        set_cell(ws, target_row, COL_X, record[LOAN_COL_L - 1])
        # Y列：放款明细 AL列 转让日
        set_cell(ws, target_row, COL_Y, record[LOAN_COL_AL - 1])
        # Z列：放款明细 AQ列 应收账款金额/价值
        set_cell(ws, target_row, COL_Z, record[LOAN_COL_AQ - 1])
        # AA列：放款明细 AR列 转让总金额
        set_cell(ws, target_row, COL_AA, record[LOAN_COL_AR - 1])
        # AB列：放款明细 AO列 原始账款到期日
        set_cell(ws, target_row, COL_AB, record[LOAN_COL_AO - 1])
        # AD列：放款明细 AM列 转让通知函编号
        set_cell(ws, target_row, COL_AD, record[LOAN_COL_AM - 1])
        # AG列：放款明细 AN列 中登登记编号
        set_cell(ws, target_row, COL_AG, record[LOAN_COL_AN - 1])
        # AH列：固定为"正常"
        set_cell(ws, target_row, COL_AH, "正常")

        added += 1
    return added


# =============================================================================
# 中登登记表 Sheet 处理函数
# =============================================================================

def append_zhongdeng_block(ws, template_cache, template_height, template_row_index: int, rows: Iterable[Sequence]) -> int:
    added = 0
    start_row = ws.max_row
    for record in rows:
        target_row = start_row + added + 1
        apply_template_row(ws, template_cache, target_row, template_row_index, template_height)

        set_cell(ws, target_row, COL_A, "=ROW()-1")
        set_cell(ws, target_row, COL_B, None)
        set_cell(ws, target_row, COL_C, record[ZD_COL_C - 1])
        set_cell(ws, target_row, COL_D, record[ZD_COL_D - 1])
        set_cell(ws, target_row, COL_F, record[ZD_COL_E - 1])
        set_cell(ws, target_row, COL_G, record[ZD_COL_F - 1])
        zh_g_value = record[ZD_COL_G - 1]
        set_cell(ws, target_row, COL_H, "" if zh_g_value is None else str(zh_g_value))
        ws.cell(row=target_row, column=COL_H).number_format = "@"
        set_cell(ws, target_row, COL_I, record[ZD_COL_H - 1])
        set_cell(ws, target_row, COL_J, record[ZD_COL_I - 1])
        set_cell(ws, target_row, COL_K, record[ZD_COL_J - 1])
        ws.cell(row=target_row, column=COL_K).alignment = Alignment(wrap_text=False, horizontal="center", vertical="center")
        set_cell(ws, target_row, COL_L, record[ZD_COL_K - 1])
        set_cell(ws, target_row, COL_M, record[ZD_COL_L - 1])
        ws.cell(row=target_row, column=COL_M).alignment = Alignment(wrap_text=False, horizontal="center", vertical="center")
        set_cell(ws, target_row, COL_N, record[ZD_COL_M - 1])
        set_cell(ws, target_row, COL_O, record[ZD_COL_N - 1])
        set_cell(ws, target_row, COL_P, record[ZD_COL_O - 1])
        set_cell(ws, target_row, COL_Q, record[ZD_COL_P - 1])
        set_cell(ws, target_row, COL_R, "/")
        set_cell(ws, target_row, COL_S, record[ZD_COL_R - 1])
        set_cell(ws, target_row, COL_T, record[ZD_COL_S - 1])
        set_cell(ws, target_row, COL_U, record[ZD_COL_T - 1])
        set_cell(ws, target_row, COL_V, record[ZD_COL_U - 1])
        set_cell(ws, target_row, COL_W, record[ZD_COL_W - 1])
        set_cell(ws, target_row, COL_X, record[ZD_COL_X - 1])
        set_cell(ws, target_row, COL_Y, record[ZD_COL_Y - 1])

        added += 1
    return added


def process_financing_repayment_sheet(wb, loan_rows: List[Sequence], factoring_repay_rows: List[Sequence],
                                       refactoring_repay_rows: List[Sequence], target_date: dt.date) -> int:
    """
    处理【融资及还款明细】sheet
    返回新增行数
    """
    ws = find_sheet_by_name(wb, SHEET_FINANCING_REPAYMENT)
    template_cache = cache_template_row(ws, TEMPLATE_ROW_INDEX)
    template_height = ws.row_dimensions[TEMPLATE_ROW_INDEX].height

    last_date = get_last_existing_date(ws)
    if last_date and target_date <= last_date:
        print(f"[融资及还款明细] 无需更新（{target_date} <= {last_date}）")
        return 0

    append_start_row = find_last_data_row(ws) + 1

    total_added = 0
    total_added += append_loan_block(ws, template_cache, template_height, TEMPLATE_ROW_INDEX, loan_rows)
    total_added += append_repay_block(ws, template_cache, template_height, TEMPLATE_ROW_INDEX, factoring_repay_rows, "保理")
    total_added += append_repay_block(ws, template_cache, template_height, TEMPLATE_ROW_INDEX, refactoring_repay_rows, "再保理")

    if total_added:
        merge_ai_with_sum(ws, append_start_row, ws.max_row, target_date)
        apply_b_column_conditional_format(ws, append_start_row, ws.max_row)

    print(f"[融资及还款明细] 新增 {total_added} 行：放款 {len(loan_rows)}，保理还款 {len(factoring_repay_rows)}，再保理还款 {len(refactoring_repay_rows)}")
    return total_added


def process_asset_detail_sheet(wb, loan_rows: List[Sequence], target_date: dt.date) -> int:
    """
    处理【资产明细】sheet
    数据来源仅为放款明细
    返回新增行数
    """
    ws = find_sheet_by_name(wb, SHEET_ASSET_DETAIL)
    template_cache = cache_template_row(ws, TEMPLATE_ROW_INDEX)
    template_height = ws.row_dimensions[TEMPLATE_ROW_INDEX].height

    total_added = append_asset_detail_block(ws, template_cache, template_height, TEMPLATE_ROW_INDEX, loan_rows)

    print(f"[资产明细] 新增 {total_added} 行")
    return total_added


def process_zhongdeng_sheet(wb, zhongdeng_rows: List[Sequence]) -> int:
    if not zhongdeng_rows:
        print("[中登登记表] 匹配到 0 行，跳过追加")
        return 0

    ws = find_sheet_by_name(wb, SHEET_ZHONGDENG)
    template_cache = cache_template_row(ws, TEMPLATE_ROW_INDEX)
    template_height = ws.row_dimensions[TEMPLATE_ROW_INDEX].height

    added = append_zhongdeng_block(ws, template_cache, template_height, TEMPLATE_ROW_INDEX, zhongdeng_rows)
    print(f"[中登登记表] 新增 {added} 行")
    return added


def process_interest_sheet(wb, factoring_interest_rows: List[Sequence], refactoring_interest_rows: List[Sequence]) -> int:
    if not factoring_interest_rows and not refactoring_interest_rows:
        print("[利息缴纳] 目标日期无资金费记录，跳过")
        return 0

    ws = find_sheet_by_name(wb, SHEET_INTEREST)
    template_cache = cache_template_row(ws, TEMPLATE_ROW_INDEX)
    template_height = ws.row_dimensions[TEMPLATE_ROW_INDEX].height

    total_added = 0
    total_added += append_interest_rows(ws, template_cache, template_height, TEMPLATE_ROW_INDEX, factoring_interest_rows, "保理")
    total_added += append_interest_rows(ws, template_cache, template_height, TEMPLATE_ROW_INDEX, refactoring_interest_rows, "再保理")

    print(f"[利息缴纳] 合计新增 {total_added} 行")
    return total_added


# =============================================================================
# 客户表 Sheet 处理函数
# =============================================================================

def collect_customer_names_from_financing(ws, target_date: dt.date) -> List[str]:
    """
    从【融资及还款明细】sheet 中筛选 W 列=目标日期的行，收集 H/I/J 列名称，按出现顺序去重
    """
    names: List[str] = []
    seen: set[str] = set()
    for row_idx in range(2, ws.max_row + 1):
        loan_date = normalize_excel_date(ws.cell(row=row_idx, column=COL_W).value)
        if loan_date != target_date:
            continue
        for col_idx in (COL_H, COL_I, COL_J):
            name = normalize_string(ws.cell(row=row_idx, column=col_idx).value)
            if not name or name == "/":
                continue
            if name not in seen:
                seen.add(name)
                names.append(name)
    return names


def collect_existing_customer_names(ws) -> set[str]:
    existing: set[str] = set()
    for row_idx in range(2, ws.max_row + 1):
        name = normalize_string(ws.cell(row=row_idx, column=COL_E).value)
        if name and name != "/":
            existing.add(name)
    return existing


def build_asset_lookup_for_customers(ws, target_names: set[str]) -> Dict[str, Dict[str, object]]:
    """
    在【资产明细】sheet 中查找 P/R 列匹配的最早行，返回通道与日期信息
    """
    if not target_names:
        return {}

    lookup: Dict[str, Dict[str, object]] = {}
    for row_idx in range(4, ws.max_row + 1):
        for col_idx in (COL_P, COL_R):
            candidate = normalize_string(ws.cell(row=row_idx, column=col_idx).value)
            if candidate and candidate in target_names and candidate not in lookup:
                lookup[candidate] = {
                    "channel": ws.cell(row=row_idx, column=COL_C).value,
                    "first_date": ws.cell(row=row_idx, column=COL_Y).value,
                }
        if len(lookup) == len(target_names):
            break
    return lookup


def get_source_cell(row: Sequence, col_idx: int):
    if row is None or col_idx <= 0:
        return None
    idx = col_idx - 1
    if idx < len(row):
        return row[idx]
    return None


def load_customer_source_map(path: Path) -> Dict[str, Sequence]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[CUSTOMER_SOURCE_SHEET] if CUSTOMER_SOURCE_SHEET in wb.sheetnames else wb.active
        mapping: Dict[str, Sequence] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = normalize_string(get_source_cell(row, CUSTOMER_SRC_COL_NAME))
            if not name or name == "/":
                continue
            if name not in mapping:
                mapping[name] = row
        return mapping
    finally:
        wb.close()


def map_channel_value(value):
    text = normalize_string(value)
    if not text:
        return None
    if text == "平台推荐":
        return "渠道合作"
    if text == "公司自拓":
        return "国富自营"
    return text


def append_customer_rows(
    ws,
    template_cache,
    template_height,
    template_row_index: int,
    names: List[str],
    asset_lookup: Dict[str, Dict[str, object]],
    customer_source: Dict[str, Sequence]
) -> tuple[int, List[str], List[str]]:
    added = 0
    missing_asset: List[str] = []
    missing_source: List[str] = []
    start_row = find_last_data_row(ws)

    for name in names:
        target_row = start_row + added + 1
        apply_template_row(ws, template_cache, target_row, template_row_index, template_height)

        asset_info = asset_lookup.get(name)
        source_row = customer_source.get(name)
        if asset_info is None:
            missing_asset.append(name)
        if source_row is None:
            missing_source.append(name)

        set_cell(ws, target_row, COL_A, "=ROW()-1")
        set_cell(ws, target_row, COL_B, map_channel_value(asset_info.get("channel") if asset_info else None))
        set_cell(ws, target_row, COL_C, asset_info.get("first_date") if asset_info else None)
        set_cell(ws, target_row, COL_D, get_source_cell(source_row, CUSTOMER_SRC_COL_REGION))
        set_cell(ws, target_row, COL_E, name)
        set_cell(ws, target_row, COL_F, get_source_cell(source_row, CUSTOMER_SRC_COL_ROLE))
        set_cell(ws, target_row, COL_G, get_source_cell(source_row, CUSTOMER_SRC_COL_CODE))
        set_cell(ws, target_row, COL_H, get_source_cell(source_row, CUSTOMER_SRC_COL_INDUSTRY))
        for col_idx in (COL_I, COL_J, COL_K, COL_L, COL_M, COL_N):
            set_cell(ws, target_row, col_idx, None)
        set_cell(ws, target_row, COL_O, get_source_cell(source_row, CUSTOMER_SRC_COL_ECONOMIC))
        set_cell(ws, target_row, COL_P, get_source_cell(source_row, CUSTOMER_SRC_COL_SCALE))
        set_cell(ws, target_row, COL_Q, get_source_cell(source_row, CUSTOMER_SRC_COL_REGISTER_ADDR))
        set_cell(ws, target_row, COL_R, get_source_cell(source_row, CUSTOMER_SRC_COL_BUSINESS_ADDR))
        set_cell(ws, target_row, COL_S, get_source_cell(source_row, CUSTOMER_SRC_COL_ROLE))
        set_cell(ws, target_row, COL_T, get_source_cell(source_row, CUSTOMER_SRC_COL_LEGAL_REP))
        set_cell(ws, target_row, COL_U, get_source_cell(source_row, CUSTOMER_SRC_COL_LEGAL_ID))
        for col_idx in (COL_V, COL_W, COL_X, COL_Y):
            set_cell(ws, target_row, col_idx, None)

        added += 1

    return added, missing_asset, missing_source


def process_customer_sheet(wb, customer_source_path: Path, target_date: dt.date) -> int:
    ws_financing = find_sheet_by_name(wb, SHEET_FINANCING_REPAYMENT)
    ws_asset = find_sheet_by_name(wb, SHEET_ASSET_DETAIL)
    ws_customer = find_sheet_by_name(wb, SHEET_CUSTOMER)

    candidate_names = collect_customer_names_from_financing(ws_financing, target_date)
    if not candidate_names:
        print("[客户表] 目标日期未发现新增客户，跳过")
        return 0

    existing_names = collect_existing_customer_names(ws_customer)
    new_names = [name for name in candidate_names if name not in existing_names]
    if not new_names:
        print("[客户表] 目标日期客户已全部存在，跳过追加")
        return 0

    asset_lookup = build_asset_lookup_for_customers(ws_asset, set(new_names))
    customer_source = load_customer_source_map(customer_source_path)

    template_cache = cache_template_row(ws_customer, TEMPLATE_ROW_INDEX)
    template_height = ws_customer.row_dimensions[TEMPLATE_ROW_INDEX].height

    added, missing_asset, missing_source = append_customer_rows(
        ws_customer,
        template_cache,
        template_height,
        TEMPLATE_ROW_INDEX,
        new_names,
        asset_lookup,
        customer_source
    )

    if added:
        print(f"[客户表] 新增 {added} 行（资产明细缺失 {len(missing_asset)}，下载客户表缺失 {len(missing_source)}）")
    else:
        print("[客户表] 未新增行")

    if missing_asset:
        print("[客户表] 警告：资产明细未找到 -> " + ", ".join(missing_asset))
    if missing_source:
        print("[客户表] 警告：下载客户表未找到 -> " + ", ".join(missing_source))

    return added


def main():
    args = parse_args()
    target_date = parse_input_date(args.date)

    ledger_path = Path(args.ledger).resolve()
    loan_path = Path(args.loan).resolve()
    factoring_path = Path(args.factoring_repay).resolve()
    refactoring_path = Path(args.refactoring_repay).resolve()
    zhongdeng_path = Path(args.zhongdeng).resolve()
    customer_path = Path(args.customer).resolve()
    output_path = Path(args.output).resolve()

    # 加载台账工作簿
    wb = load_workbook(ledger_path, data_only=False)

    # 收集数据（统一查询条件）
    loan_rows = collect_loan_rows(loan_path, target_date)
    factoring_repay_rows = collect_repay_rows(factoring_path, target_date, fee_type="本金")
    refactoring_repay_rows = collect_repay_rows(refactoring_path, target_date, fee_type="本金")
    factoring_interest_rows = collect_repay_rows(factoring_path, target_date, fee_type="资金费")
    refactoring_interest_rows = collect_repay_rows(refactoring_path, target_date, fee_type="资金费")
    finance_codes = set()
    for row in loan_rows:
        code = normalize_string(row[LOAN_COL_L - 1])
        if code:
            finance_codes.add(code)
    zhongdeng_rows = collect_zhongdeng_rows(zhongdeng_path, finance_codes)

    # 处理各个 sheet
    total_added = 0
    total_added += process_financing_repayment_sheet(wb, loan_rows, factoring_repay_rows, refactoring_repay_rows, target_date)
    total_added += process_asset_detail_sheet(wb, loan_rows, target_date)
    total_added += process_zhongdeng_sheet(wb, zhongdeng_rows)
    total_added += process_customer_sheet(wb, customer_path, target_date)
    total_added += process_interest_sheet(wb, factoring_interest_rows, refactoring_interest_rows)

    # 保存输出
    wb.save(output_path)
    print(f"[ledger_daily] 完成写入 -> {output_path}，总计新增 {total_added} 行")


if __name__ == "__main__":
    main()
